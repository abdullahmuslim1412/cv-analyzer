"""
AI-Powered CV Analyzer — Exponentiq
-------------------------------------
Structured, evidence-based candidate screening for HR teams.
Upload a job description and a batch of CVs; each candidate is scored on
seven weighted dimensions with a written rationale, ranked, and laid out
side-by-side against the source CV for a human reviewer to audit.

Run:
    pip install -r requirements.txt
    streamlit run app.py

Configuration (set by the workspace admin, never shown in the UI):
    .streamlit/secrets.toml
        OPENROUTER_API_KEY = "sk-or-..."
        OPENROUTER_MODEL   = "openai/gpt-4o-mini"   # optional override
"""

import io
import json
import os
import re

import fitz  # PyMuPDF
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openai import OpenAI

# ─────────────────────────────────────────────
# BRAND TOKENS
# ─────────────────────────────────────────────
INK = "#14213D"        # primary text / header
SLATE = "#5B6472"      # secondary text
FOG = "#F4F5F7"        # page background
PAPER = "#FFFFFF"      # card background
LINE = "#E3E6EB"       # hairline borders
TEAL = "#0F7173"       # accent / CTA
GREEN = "#1E8E5A"
GREEN_BG = "#E7F5EC"
AMBER = "#B7791F"
AMBER_BG = "#FDF3DF"
RED = "#C0392B"
RED_BG = "#FBEAE8"
LOGO_BG = "#071020"    # sampled from the Exponentiq logo asset's own background

APP_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH_CANDIDATES = [
    os.path.join(APP_DIR, "assets", "logo.png"),
    os.path.join(APP_DIR, "assets", "logo.svg"),
]

DEFAULT_MODEL = "openai/gpt-4o-mini"

st.set_page_config(page_title="AI-Powered CV Analyzer | Exponentiq", layout="wide")

DIMENSIONS = [
    "education",
    "work_experience",
    "skills",
    "certifications_and_courses",
    "languages",
    "technology_knowledge",
    "general_assessment",
]

CANDIDATE_STRUCTURE = {
    "full_name": "", "email": "", "phone": "", "location": "",
    "latest_qualification": "", "last_school_or_university": "",
    "graduation_year": "", "latest_job_title": "", "latest_employer": "",
    "latest_job_duration": "", "years_of_experience": "",
    "top_skills": [], "certifications": [], "languages": [], "systems_known": [],
}

POSITION_STRUCTURE = {
    "JobTitle": "", "Department": "", "ReportsTo": "", "JobPurpose": "",
    "KeyResponsibilities": [], "WorkModality": "", "AdditionalDetails": "",
    "Qualifications": {
        "education": [], "work_experience": [], "skills": [],
        "certifications_and_courses": [],
        "languages": [{"language": "", "proficiency": ""}],
        "systems_knowledge": [],
    },
}

EVALUATION_STRUCTURE = {
    "dimensions": [{"name": d, "value": 0, "reasoning": ""} for d in DIMENSIONS]
}


# ─────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────
def inject_css():
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600&family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@500;600&display=swap');

    html, body, [class*="css"] {{
        font-family: 'Inter', -apple-system, sans-serif;
    }}
    .stApp {{ background-color: {FOG}; }}
    #MainMenu, footer {{ visibility: hidden; }}

    /* ---- header ---- */
    .eq-header {{
        display: flex; align-items: center; gap: 18px;
        padding: 22px 4px 20px 4px;
        border-bottom: 1px solid {LINE};
        margin-bottom: 28px;
    }}
    .eq-logo-fallback {{
        width: 44px; height: 44px; border-radius: 9px;
        background: {INK};
        color: white; font-family: 'JetBrains Mono', monospace; font-weight: 600;
        display: flex; align-items: center; justify-content: center;
        font-size: 15px; letter-spacing: 0.5px; flex-shrink: 0;
    }}
    .eq-logo-chip {{
        background: {LOGO_BG}; border-radius: 10px;
        padding: 10px 18px; display: inline-flex; align-items: center; flex-shrink: 0;
    }}
    .eq-header-divider {{
        width: 1px; align-self: stretch; background: {LINE}; margin: 2px 0;
    }}
    .eq-title {{
        font-family: 'Fraunces', serif; font-weight: 600; font-size: 1.65rem;
        color: {INK}; line-height: 1.15; margin: 0;
    }}
    .eq-subtitle {{
        font-family: 'Inter', sans-serif; font-size: 0.85rem; color: {SLATE};
        margin-top: 3px; letter-spacing: 0.1px;
    }}
    .eq-brandline {{
        margin-left: auto; text-align: right; font-family: 'Inter', sans-serif;
        font-size: 0.72rem; color: {SLATE}; text-transform: uppercase; letter-spacing: 1px;
    }}

    /* ---- KPI cards ---- */
    .eq-kpi {{
        background: {PAPER}; border: 1px solid {LINE}; border-radius: 10px;
        padding: 16px 18px; height: 100%;
    }}
    .eq-kpi-label {{
        font-size: 0.72rem; color: {SLATE}; text-transform: uppercase;
        letter-spacing: 0.8px; font-weight: 600; margin-bottom: 6px;
    }}
    .eq-kpi-value {{
        font-family: 'JetBrains Mono', monospace; font-weight: 600; font-size: 1.55rem;
        color: {INK};
    }}
    .eq-kpi-sub {{
        font-size: 0.76rem; color: {SLATE}; margin-top: 2px;
    }}

    /* ---- section labels ---- */
    .eq-section-label {{
        font-size: 0.78rem; font-weight: 600; color: {INK};
        text-transform: uppercase; letter-spacing: 0.6px;
        margin-bottom: 10px; padding-bottom: 8px; border-bottom: 1px solid {LINE};
    }}

    /* ---- score ledger card (right panel) ---- */
    .eq-score-row {{
        border-left: 3px solid var(--tier-color, {LINE});
        background: {PAPER}; border-radius: 6px;
        padding: 11px 14px; margin-bottom: 9px;
        border-top: 1px solid {LINE}; border-right: 1px solid {LINE}; border-bottom: 1px solid {LINE};
    }}
    .eq-score-row-head {{
        display: flex; justify-content: space-between; align-items: baseline;
    }}
    .eq-score-name {{ font-weight: 600; font-size: 0.88rem; color: {INK}; }}
    .eq-score-value {{
        font-family: 'JetBrains Mono', monospace; font-weight: 600; font-size: 0.95rem;
    }}
    .eq-score-reasoning {{
        font-size: 0.82rem; color: {SLATE}; margin-top: 4px; line-height: 1.45;
    }}

    /* ---- cv text panel ---- */
    .eq-cv-panel {{
        height: 640px; overflow-y: auto; border: 1px solid {LINE}; border-radius: 8px;
        padding: 18px; background: {PAPER}; font-size: 0.83rem; white-space: pre-wrap;
        line-height: 1.55; color: #2A2F38; font-family: 'Inter', sans-serif;
    }}
    .eq-score-panel-wrap {{ height: 640px; overflow-y: auto; padding-right: 4px; }}

    /* ---- rank pill in detail selector row ---- */
    .eq-overall-badge {{
        font-family: 'JetBrains Mono', monospace; font-weight: 600; font-size: 1.1rem;
        padding: 3px 12px; border-radius: 20px; display: inline-block;
    }}

    /* buttons */
    .stButton > button {{
        background-color: {INK}; color: white; border: none; border-radius: 7px;
        font-weight: 600; font-size: 0.85rem;
    }}
    .stButton > button:hover {{ background-color: {TEAL}; color: white; }}
    .stDownloadButton > button {{
        border: 1px solid {INK}; color: {INK}; background: {PAPER}; border-radius: 7px; font-weight: 600;
    }}
    </style>
    """, unsafe_allow_html=True)


def render_header():
    logo_html = None
    for path in LOGO_PATH_CANDIDATES:
        if os.path.exists(path):
            import base64
            with open(path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode()
            ext = "svg+xml" if path.endswith(".svg") else "png"
            logo_html = (
                f'<div class="eq-logo-chip">'
                f'<img src="data:image/{ext};base64,{b64}" style="height:30px; width:auto; display:block;" />'
                f'</div>'
            )
            break
    if logo_html is None:
        logo_html = '<div class="eq-logo-fallback">EQ</div>'

    st.markdown(f"""
    <div class="eq-header">
        {logo_html}
        <div class="eq-header-divider"></div>
        <div>
            <p class="eq-title">AI-Powered CV Analyzer</p>
            <p class="eq-subtitle">Structured, evidence-based candidate screening</p>
        </div>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# NUMBER HANDLING (robust against LLM output like "8/10", "N/A", None)
# ─────────────────────────────────────────────
def safe_float(v, default=0.0):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except (TypeError, ValueError):
            return default
    if isinstance(v, str):
        match = re.search(r"-?\d+(\.\d+)?", v)
        if match:
            try:
                return float(match.group(0))
            except ValueError:
                return default
    return default


def fmt_score(v):
    v = safe_float(v)
    return f"{v:g}"


# ─────────────────────────────────────────────
# PDF / LLM HELPERS
# ─────────────────────────────────────────────
def extract_pdf_text(uploaded_file) -> str:
    if uploaded_file is None:
        raise ValueError("No file was provided to extract text from.")

    if hasattr(uploaded_file, "getvalue"):
        data = uploaded_file.getvalue()
    elif hasattr(uploaded_file, "read"):
        data = uploaded_file.read()
    else:
        raise ValueError(
            f"Unsupported file object of type {type(uploaded_file)!r} — expected an uploaded PDF file."
        )

    if not data:
        raise ValueError("The uploaded file appears to be empty.")

    doc = fitz.open(stream=data, filetype="pdf")
    try:
        text = "".join(page.get_text() for page in doc)
    finally:
        doc.close()

    if not text.strip():
        raise ValueError(
            "No extractable text found in this PDF (it may be a scanned/image-only PDF)."
        )
    return text


def extract_json(raw: str) -> dict:
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip()
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if match:
        raw = match.group(0)
    return json.loads(raw)


def call_llm(client, model, system_prompt, user_prompt) -> str:
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    )
    return response.choices[0].message.content.strip()


def normalize_jd(client, model, jd_text):
    raw = call_llm(
        client, model,
        "You are a recruiter assistant. Convert job descriptions into structured JSON. Return ONLY valid JSON, no explanation.",
        f"Extract and return this job description as JSON using exactly this structure:\n\n{json.dumps(POSITION_STRUCTURE, indent=2)}\n\nJob Description:\n{jd_text}",
    )
    return extract_json(raw)


def extract_profile(client, model, cv_text):
    raw = call_llm(
        client, model,
        "You are a CV parser. Extract candidate information into structured JSON. Return ONLY valid JSON, no explanation.",
        f"Extract candidate details from this CV using exactly this JSON structure:\n\n{json.dumps(CANDIDATE_STRUCTURE, indent=2)}\n\nCV:\n{cv_text}",
    )
    return extract_json(raw)


def evaluate_cv(client, model, structured_jd, cv_text):
    raw = call_llm(
        client, model,
        "You are a professional recruiter. Evaluate candidates against job descriptions strictly and honestly. "
        "Scores must be plain numbers from 0 to 10 (not strings like '8/10'). Return ONLY valid JSON, no explanation.",
        f"""Evaluate this candidate's CV against the job description.
Use this exact JSON structure (scores are plain numbers 0-10):
{json.dumps(EVALUATION_STRUCTURE, indent=2)}

Job Description (structured):
{json.dumps(structured_jd, indent=2)}

Candidate CV:
{cv_text}

Return valid JSON only.""",
    )
    return extract_json(raw)


def score_tier(v):
    v = safe_float(v)
    if v >= 7:
        return GREEN, GREEN_BG
    elif v >= 4:
        return AMBER, AMBER_BG
    return RED, RED_BG


def score_row_html(name, value, reasoning):
    fg, bg = score_tier(value)
    label = name.replace("_", " ").title()
    reasoning = reasoning or "No rationale provided."
    return f"""
    <div class="eq-score-row" style="--tier-color:{fg};">
        <div class="eq-score-row-head">
            <span class="eq-score-name">{label}</span>
            <span class="eq-score-value" style="color:{fg};">{fmt_score(value)}/10</span>
        </div>
        <div class="eq-score-reasoning">{reasoning}</div>
    </div>
    """


SCORE_COLORSCALE = [(0.0, RED), (0.5, AMBER), (1.0, GREEN)]


def build_excel(results_df: pd.DataFrame) -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        results_df.to_excel(writer, index=False, sheet_name="Evaluation Results")
        ws = writer.sheets["Evaluation Results"]

        header_fill = PatternFill("solid", fgColor="14213D")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                              top=Side(style="thin"), bottom=Side(style="thin"))

        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center_align, thin_border

        score_cols = [c for c in results_df.columns if c.endswith("_score")]
        score_col_idx = [results_df.columns.get_loc(c) + 1 for c in score_cols]

        tier_fills = {
            "green": PatternFill("solid", fgColor="E7F5EC"),
            "amber": PatternFill("solid", fgColor="FDF3DF"),
            "red": PatternFill("solid", fgColor="FBEAE8"),
        }

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = thin_border
                if cell.column in score_col_idx:
                    v = safe_float(cell.value)
                    if v >= 7:
                        cell.fill = tier_fills["green"]
                    elif v >= 4:
                        cell.fill = tier_fills["amber"]
                    else:
                        cell.fill = tier_fills["red"]

        for col_num in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_num)
            max_len = max((len(str(ws.cell(row=r, column=col_num).value or ""))
                           for r in range(1, ws.max_row + 1)), default=10)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)
        ws.freeze_panes = "A2"

    return buf.getvalue()


# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
if "results" not in st.session_state:
    st.session_state.results = []
if "jd_json" not in st.session_state:
    st.session_state.jd_json = None
if "jd_text" not in st.session_state:
    st.session_state.jd_text = ""

inject_css()
render_header()


def _get_secret(key, default=None):
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default


api_key = _get_secret("OPENROUTER_API_KEY", "")
model = _get_secret("OPENROUTER_MODEL", DEFAULT_MODEL)

# ─────────────────────────────────────────────
# SIDEBAR — uploads only, no technical controls
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("**Job description**")
    jd_file = st.file_uploader("Job description (PDF)", type="pdf", label_visibility="collapsed")

    st.markdown("**Candidate CVs**")
    cv_files = st.file_uploader("Candidate CVs (PDF)", type="pdf", accept_multiple_files=True,
                                 label_visibility="collapsed")

    st.write("")
    run = st.button("Run analysis", type="primary", width='stretch',
                     disabled=not (api_key and jd_file and cv_files))

    if not api_key:
        st.caption("Analysis is not yet configured for this workspace. Contact your administrator.")
    elif not (jd_file and cv_files):
        st.caption("Upload a job description and at least one CV to begin.")

# ─────────────────────────────────────────────
# RUN ANALYSIS
# ─────────────────────────────────────────────
if run:
    if not jd_file or not cv_files:
        st.error("Upload both a job description and at least one CV, then run analysis again.")
        st.stop()

    client = OpenAI(api_key=api_key, base_url="https://openrouter.ai/api/v1")
    st.session_state.results = []

    with st.spinner("Reading job description..."):
        try:
            jd_text = extract_pdf_text(jd_file)
        except Exception as e:
            st.error(f"Could not read the job description PDF: {e}")
            st.stop()
        st.session_state.jd_text = jd_text
        try:
            st.session_state.jd_json = normalize_jd(client, model, jd_text)
        except Exception as e:
            st.error(f"Could not interpret the job description: {e}")
            st.session_state.jd_json = POSITION_STRUCTURE

    progress = st.progress(0.0, text="Starting candidate analysis...")
    n = len(cv_files)
    for i, cv_file in enumerate(cv_files):
        progress.progress(i / n, text=f"Analyzing {cv_file.name}...")

        try:
            cv_text = extract_pdf_text(cv_file)
        except Exception as e:
            st.warning(f"Skipped {cv_file.name} — {e}")
            continue

        try:
            profile = extract_profile(client, model, cv_text)
        except Exception as e:
            st.warning(f"Could not extract profile details for {cv_file.name}: {e}")
            profile = CANDIDATE_STRUCTURE.copy()

        try:
            eval_data = evaluate_cv(client, model, st.session_state.jd_json, cv_text)
        except Exception as e:
            st.warning(f"Could not score {cv_file.name}: {e}")
            eval_data = EVALUATION_STRUCTURE.copy()

        dims = {d["name"]: d for d in eval_data.get("dimensions", [])}
        scores = [safe_float(dims.get(d, {}).get("value", 0)) for d in DIMENSIONS]
        overall = round(sum(scores) / len(scores), 2) if scores else 0.0

        st.session_state.results.append({
            "filename": cv_file.name,
            "cv_text": cv_text,
            "profile": profile,
            "dimensions": dims,
            "overall_score": overall,
        })

    progress.progress(1.0, text="Done")
    if not st.session_state.results:
        st.error("No candidates could be analyzed. Check the uploaded files and try again.")
        st.stop()
    st.success(f"Analyzed {len(st.session_state.results)} of {n} candidate(s).")

# ─────────────────────────────────────────────
# EMPTY STATE
# ─────────────────────────────────────────────
if not st.session_state.results:
    st.markdown(f"""
    <div style="border:1px dashed {LINE}; border-radius:10px; padding:48px; text-align:center; background:{PAPER};">
        <p style="font-family:'Fraunces',serif; font-size:1.2rem; color:{INK}; margin-bottom:6px;">
            No candidates analyzed yet
        </p>
        <p style="color:{SLATE}; font-size:0.9rem;">
            Upload a job description and one or more CVs in the sidebar, then run analysis.
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────
# BUILD RESULTS TABLE
# ─────────────────────────────────────────────
results = st.session_state.results
df_rows = []
for r in results:
    row = {
        "Candidate": r["profile"].get("full_name") or r["filename"],
        "Filename": r["filename"],
        "Overall Score": safe_float(r["overall_score"]),
    }
    for d in DIMENSIONS:
        row[d.replace("_", " ").title()] = safe_float(r["dimensions"].get(d, {}).get("value", 0))
    df_rows.append(row)

df = pd.DataFrame(df_rows).sort_values("Overall Score", ascending=False).reset_index(drop=True)
df.insert(0, "Rank", df.index + 1)

# ─────────────────────────────────────────────
# KPI STRIP
# ─────────────────────────────────────────────
avg_score = round(df["Overall Score"].mean(), 1)
top_row = df.iloc[0]
strong_matches = int((df["Overall Score"] >= 7).sum())

k1, k2, k3, k4 = st.columns(4)
for col, label, value, sub in [
    (k1, "Candidates evaluated", str(len(df)), ""),
    (k2, "Top candidate", top_row["Candidate"], f"{fmt_score(top_row['Overall Score'])}/10"),
    (k3, "Average score", f"{fmt_score(avg_score)}/10", ""),
    (k4, "Strong matches (≥7)", str(strong_matches), f"of {len(df)} candidates"),
]:
    with col:
        st.markdown(f"""
        <div class="eq-kpi">
            <div class="eq-kpi-label">{label}</div>
            <div class="eq-kpi-value">{value}</div>
            <div class="eq-kpi-sub">{sub}</div>
        </div>
        """, unsafe_allow_html=True)

st.write("")
tab_overview, tab_detail = st.tabs(["Overview", "Candidate detail"])

# ---------- OVERVIEW TAB ----------
with tab_overview:
    col1, col2 = st.columns([1.3, 1])

    with col1:
        st.markdown('<div class="eq-section-label">Ranked results</div>', unsafe_allow_html=True)
        score_cols = [c for c in df.columns if c not in ("Rank", "Candidate", "Filename")]

        def _highlight_score(v):
            fg, bg = score_tier(v)
            return f"background-color: {bg}; color: {fg}; font-weight: 600;"

        try:
            styler = df.style
            style_fn = styler.map if hasattr(styler, "map") else styler.applymap
            styled = style_fn(_highlight_score, subset=score_cols)
            st.dataframe(styled, width='stretch', height=min(60 + 40 * len(df), 500))
        except Exception:
            st.dataframe(df, width='stretch', height=min(60 + 40 * len(df), 500))

        excel_bytes = build_excel(df.drop(columns=["Filename"]))
        st.download_button("Download Excel report", data=excel_bytes,
                            file_name="CV_Evaluation_Results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with col2:
        st.markdown('<div class="eq-section-label">Overall score by candidate</div>', unsafe_allow_html=True)
        try:
            fig_bar = px.bar(df, x="Overall Score", y="Candidate", orientation="h",
                              color="Overall Score", color_continuous_scale=SCORE_COLORSCALE,
                              range_color=[0, 10])
            fig_bar.update_layout(
                yaxis={"categoryorder": "total ascending"}, height=400,
                plot_bgcolor=PAPER, paper_bgcolor=PAPER,
                font=dict(family="Inter, sans-serif", color=INK),
                margin=dict(l=0, r=10, t=10, b=10),
            )
            st.plotly_chart(fig_bar, width='stretch')
        except Exception as e:
            st.warning(f"Could not render the score chart: {e}")

    st.markdown('<div class="eq-section-label" style="margin-top:14px;">Dimension breakdown across candidates</div>',
                unsafe_allow_html=True)
    try:
        dim_cols = [d.replace("_", " ").title() for d in DIMENSIONS]
        heat_df = df.set_index("Candidate")[dim_cols]
        fig_heat = px.imshow(heat_df, text_auto=True, aspect="auto",
                              color_continuous_scale=SCORE_COLORSCALE, zmin=0, zmax=10)
        fig_heat.update_layout(
            height=120 + 40 * len(df),
            plot_bgcolor=PAPER, paper_bgcolor=PAPER,
            font=dict(family="Inter, sans-serif", color=INK),
            margin=dict(l=0, r=10, t=10, b=10),
        )
        st.plotly_chart(fig_heat, width='stretch')
    except Exception as e:
        st.warning(f"Could not render the dimension breakdown: {e}")

# ---------- DETAIL TAB ----------
with tab_detail:
    names = [r["profile"].get("full_name") or r["filename"] for r in results]
    choice = st.selectbox("Select candidate", names, label_visibility="collapsed")
    candidate = next(r for r, nm in zip(results, names) if nm == choice)
    profile = candidate["profile"]
    dims = candidate["dimensions"]
    overall = safe_float(candidate["overall_score"])
    fg, bg = score_tier(overall)

    hcol1, hcol2 = st.columns([3, 1])
    with hcol1:
        st.markdown(f"""<p style="font-family:'Fraunces',serif; font-size:1.3rem; color:{INK}; margin-bottom:0;">{choice}</p>""",
                     unsafe_allow_html=True)
    with hcol2:
        st.markdown(f"""
        <div style="text-align:right;">
            <span class="eq-overall-badge" style="background:{bg}; color:{fg};">{fmt_score(overall)}/10</span>
        </div>
        """, unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Email", profile.get("email") or "—")
    c2.metric("Location", profile.get("location") or "—")
    c3.metric("Latest role", profile.get("latest_job_title") or "—")
    c4.metric("Experience", str(profile.get("years_of_experience") or "—"))

    try:
        radar_vals = [safe_float(dims.get(d, {}).get("value", 0)) for d in DIMENSIONS]
        radar_labels = [d.replace("_", " ").title() for d in DIMENSIONS]
        fig_radar = go.Figure()
        fig_radar.add_trace(go.Scatterpolar(
            r=radar_vals + [radar_vals[0]],
            theta=radar_labels + [radar_labels[0]],
            fill="toself", name=choice,
            line=dict(color=TEAL, width=2),
            fillcolor="rgba(15,113,115,0.15)",
        ))
        fig_radar.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 10], gridcolor=LINE),
                       angularaxis=dict(gridcolor=LINE)),
            showlegend=False, height=360, margin=dict(t=20, b=20),
            plot_bgcolor=PAPER, paper_bgcolor=PAPER,
            font=dict(family="Inter, sans-serif", color=INK),
        )
        st.plotly_chart(fig_radar, width='stretch')
    except Exception as e:
        st.warning(f"Could not render the score profile: {e}")

    st.markdown('<div class="eq-section-label" style="margin-top:6px;">Candidate CV and AI evaluation</div>',
                unsafe_allow_html=True)
    left, right = st.columns([1, 1])

    with left:
        st.markdown(f'<div class="eq-cv-panel">{candidate["cv_text"]}</div>', unsafe_allow_html=True)

    with right:
        rows_html = "".join(
            score_row_html(d, dims.get(d, {}).get("value", 0), dims.get(d, {}).get("reasoning", ""))
            for d in DIMENSIONS
        )
        st.markdown(f'<div class="eq-score-panel-wrap">{rows_html}</div>', unsafe_allow_html=True)
